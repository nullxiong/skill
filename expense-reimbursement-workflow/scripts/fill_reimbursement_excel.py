#!/usr/bin/env python3
import argparse
import json
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment


def clear_range(ws, cells):
    for cell_ref in cells:
        ws[cell_ref] = None


def write_trip(ws, start_row, trip):
    ws[f"B{start_row}"] = trip.get("start", "")
    ws[f"F{start_row}"] = trip.get("destination", "")
    ws[f"K{start_row}"] = trip.get("days", "")
    ws[f"O{start_row}"] = trip.get("date_range", "")

    train_row = start_row + 2
    taxi_row = start_row + 3
    self_drive_row = start_row + 4

    clear_range(
        ws,
        [
            f"C{train_row}", f"E{train_row}", f"F{train_row}", f"H{train_row}", f"K{train_row}",
            f"C{taxi_row}", f"E{taxi_row}", f"F{taxi_row}", f"H{taxi_row}", f"K{taxi_row}",
            f"C{self_drive_row}", f"E{self_drive_row}", f"F{self_drive_row}", f"H{self_drive_row}", f"K{self_drive_row}",
        ],
    )

    ws[f"C{train_row}"] = round(float(trip.get("train") or 0), 2)
    ws[f"E{train_row}"] = round(float(trip.get("hotel") or 0), 2)
    ws[f"F{train_row}"] = round(float(trip.get("subsidy") or (float(trip.get("days") or 0) * 75)), 2)
    ws[f"H{train_row}"] = trip.get("invoice_no", "")
    ws[f"K{train_row}"] = trip.get("reason", "")
    ws[f"C{taxi_row}"] = round(float(trip.get("taxi") or 0), 2)

    for row in (train_row, taxi_row, self_drive_row):
        ws[f"P{row}"] = f"=SUM(C{row}:G{row})"
    ws[f"P{start_row + 5}"] = f"=SUM(P{train_row}:P{self_drive_row})"

    for row in (train_row, taxi_row):
        ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True)
    parser.add_argument("--trips-json", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--sheet-name", default="销售费用报销表")
    parser.add_argument("--block-start-rows", default="16,22,28,34")
    parser.add_argument("--output-name")
    return parser.parse_args()


def main():
    args = parse_args()
    template = Path(args.template).expanduser()
    output_dir = Path(args.output_dir).expanduser()
    trips = json.loads(Path(args.trips_json).read_text(encoding="utf-8"))
    block_rows = [int(x.strip()) for x in args.block_start_rows.split(",") if x.strip()]

    wb = openpyxl.load_workbook(template)
    ws = wb[args.sheet_name]
    ws["P2"] = datetime.now()
    ws["P2"].number_format = "yyyy/m/d"

    for row, trip in zip(block_rows, trips):
        write_trip(ws, row, trip)

    output_dir.mkdir(parents=True, exist_ok=True)
    name = args.output_name or f"差旅报销-{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    output = output_dir / name
    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
